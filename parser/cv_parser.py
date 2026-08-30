

import os
import re
import time
import logging
import threading

# PDF handling
try:
    import PyPDF2
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

# DOCX handling
try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False


import openpyxl


try:
    import pytesseract
    from pdf2image import convert_from_path
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

logger = logging.getLogger('ai_recruitment')



TESSERACT_CMD = os.environ.get('TESSERACT_CMD', 'tesseract')
if OCR_AVAILABLE:
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

POPPLER_PATH = os.environ.get('POPPLER_PATH', None)

CV_KEYWORDS_XLSX = os.environ.get(
    'CV_KEYWORDS_XLSX',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cv_keywords.xlsx')
)

try:
    from django.conf import settings as _django_settings
    if getattr(_django_settings, 'CV_KEYWORDS_XLSX', None):
        CV_KEYWORDS_XLSX = str(_django_settings.CV_KEYWORDS_XLSX)
except Exception:

    pass


KEYWORDS_CACHE_TTL = int(os.environ.get('CV_KEYWORDS_CACHE_TTL', '300'))

_keywords_cache = {
    'skills': [],
    'education': [],
    'experience': [],
    'loaded_at': 0.0,
    'source': None,
}
_keywords_lock = threading.Lock()

_FALLBACK_SKILLS = ['python', 'java', 'javascript', 'sql', 'aws', 'docker']
_FALLBACK_EDUCATION = ['bachelor', 'master', 'phd', 'degree', 'university']
_FALLBACK_EXPERIENCE = ['experience', 'worked', 'managed', 'developed']


def _read_single_column_sheet(ws):

    values = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or row[0] is None:
            continue
        keyword = str(row[0]).strip()
        if not keyword:
            continue

        active_flag = row[-1] if len(row) > 1 else None
        if active_flag is not None:
            active_str = str(active_flag).strip().lower()
            if active_str in ('no', 'false', '0', 'n'):
                continue

        values.append(keyword.lower())
    return values


def load_keywords_from_excel(file_path=None, force_reload=False):

    path = file_path or CV_KEYWORDS_XLSX

    with _keywords_lock:
        cache_is_fresh = (
            not force_reload
            and _keywords_cache['source'] == path
            and (time.time() - _keywords_cache['loaded_at']) < KEYWORDS_CACHE_TTL
            and _keywords_cache['skills']
        )
        if cache_is_fresh:
            return (
                _keywords_cache['skills'],
                _keywords_cache['education'],
                _keywords_cache['experience'],
            )

        if not os.path.exists(path):
            logger.error(f"Keyword workbook not found at {path}. Using fallback keyword lists.")
            skills, education, experience = _FALLBACK_SKILLS, _FALLBACK_EDUCATION, _FALLBACK_EXPERIENCE
        else:
            try:
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

                skills = _read_single_column_sheet(wb['Skills']) if 'Skills' in wb.sheetnames else []
                education = _read_single_column_sheet(wb['EducationKeywords']) if 'EducationKeywords' in wb.sheetnames else []
                experience = _read_single_column_sheet(wb['ExperienceKeywords']) if 'ExperienceKeywords' in wb.sheetnames else []

                wb.close()

                if not skills:
                    logger.warning("Skills sheet was empty; falling back to default skill list.")
                    skills = _FALLBACK_SKILLS
                if not education:
                    education = _FALLBACK_EDUCATION
                if not experience:
                    experience = _FALLBACK_EXPERIENCE

                logger.info(
                    f"Loaded keyword sheet from {path}: "
                    f"{len(skills)} skills, {len(education)} education keywords, "
                    f"{len(experience)} experience keywords"
                )
            except Exception as e:
                logger.error(f"Error reading keyword workbook {path}: {e}. Using fallback keyword lists.")
                skills, education, experience = _FALLBACK_SKILLS, _FALLBACK_EDUCATION, _FALLBACK_EXPERIENCE

        _keywords_cache['skills'] = skills
        _keywords_cache['education'] = education
        _keywords_cache['experience'] = experience
        _keywords_cache['loaded_at'] = time.time()
        _keywords_cache['source'] = path

        return skills, education, experience


def _extract_pdf_text_pdfplumber(file_path):

    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if page_text:
                text += page_text + "\n"
    return text


def _extract_pdf_text_pypdf2(file_path):

    text = ""
    with open(file_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_pdf_text(file_path):

    text = ""

    if PDFPLUMBER_AVAILABLE:
        try:
            text = _extract_pdf_text_pdfplumber(file_path)
            logger.info(f"Extracted {len(text.strip())} characters from PDF using pdfplumber")
        except Exception as e:
            logger.error(f"pdfplumber read failed, falling back to PyPDF2: {e}")
            text = ""

    if len(text.strip()) < 100 and PYPDF2_AVAILABLE:
        try:
            pypdf2_text = _extract_pdf_text_pypdf2(file_path)
            if len(pypdf2_text.strip()) > len(text.strip()):
                text = pypdf2_text
                logger.info(f"Extracted {len(text.strip())} characters from PDF using PyPDF2 fallback")
        except Exception as e:
            logger.error(f"PyPDF2 read failed: {e}")

    if not PDFPLUMBER_AVAILABLE and not PYPDF2_AVAILABLE:
        logger.warning("Neither pdfplumber nor PyPDF2 available")
        return text

    # OCR fallback for scanned PDFs
    if len(text.strip()) < 100 and OCR_AVAILABLE:
        logger.info(f"OCR Processing for scanned PDF: {file_path}")
        try:
            if POPPLER_PATH:
                images = convert_from_path(file_path, poppler_path=POPPLER_PATH)
            else:
                images = convert_from_path(file_path)

            for image in images:
                ocr_text = pytesseract.image_to_string(image)
                text += ocr_text + "\n"

            logger.info(f"Extracted {len(text)} characters from PDF using OCR")
        except Exception as ocr_error:
            logger.error(f"OCR failed: {ocr_error}")

    return text


def extract_docx_text(file_path):
    """Extract text from DOCX file"""
    text = ""
    if not DOCX_AVAILABLE:
        return text
    try:
        doc = docx.Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
        logger.info(f"Extracted {len(text)} characters from DOCX")
    except Exception as e:
        logger.error(f"Error reading DOCX {file_path}: {e}")

    return text


def extract_text(file_path):
    """Main text extraction function - supports PDF and DOCX"""
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        return ""

    file_path_lower = file_path.lower()

    if file_path_lower.endswith(".pdf"):
        return extract_pdf_text(file_path)
    elif file_path_lower.endswith(".docx"):
        return extract_docx_text(file_path)
    elif file_path_lower.endswith(".doc"):

        logger.warning(f".doc format has limited support. Attempting text extraction for: {file_path}")
        try:
            # Try as if it's a docx (sometimes works)
            return extract_docx_text(file_path)
        except:
            return ""
    elif file_path_lower.endswith((".txt", ".rtf")):
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            logger.error(f"Error reading text file: {e}")
            return ""
    else:
        logger.warning(f"Unsupported file format: {file_path}")
        return ""


def extract_email(text):
    """Extract email addresses from text"""
    if not text:
        return ""


    text = text.replace("mailto:", "")
    text = re.sub(r"\s*@\s*", "@", text)
    text = re.sub(r"\s*\.\s*", ".", text)

    # Email regex pattern
    email_pattern = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
    emails = re.findall(email_pattern, text)

    if emails:
        return emails[0]  
    return ""


def extract_skills(text, skills_list=None):

    if not text:
        return ""

    if skills_list is None:
        skills_list, _, _ = load_keywords_from_excel()

    text_lower = text.lower()
    found_skills = []

    for skill in skills_list:

        pattern = r'\b' + re.escape(skill) + r'\b'
        if re.search(pattern, text_lower):
            found_skills.append(skill)

    return ", ".join(found_skills) if found_skills else ""


def extract_education(text, education_keywords=None):

    if not text:
        return ""

    if education_keywords is None:
        _, education_keywords, _ = load_keywords_from_excel()

    lines = text.split('\n')
    education_lines = []

    for line in lines:
        line_lower = line.lower().strip()
        if any(keyword in line_lower for keyword in education_keywords):
            # Clean the line
            clean_line = line.strip()
            if len(clean_line) > 10:  # Filter out very short matches
                education_lines.append(clean_line)

    return "\n".join(education_lines[:10]) if education_lines else ""


def extract_experience(text, experience_keywords=None):

    if not text:
        return ""

    if experience_keywords is None:
        _, _, experience_keywords = load_keywords_from_excel()

    lines = text.split('\n')
    experience_lines = []

    for line in lines:
        line_lower = line.lower().strip()
        if any(keyword in line_lower for keyword in experience_keywords):
            clean_line = line.strip()
            if len(clean_line) > 10:
                experience_lines.append(clean_line)

    return "\n".join(experience_lines[:15]) if experience_lines else ""


def parse_cv(file_path):

    result = {
        'text': '',
        'email': '',
        'skills': '',
        'education': '',
        'experience': '',
        'success': False,
        'error': ''
    }

    try:

        text = extract_text(file_path)
        result['text'] = text

        if not text.strip():
            result['error'] = 'No text could be extracted from the file'
            return result

        skills_list, education_keywords, experience_keywords = load_keywords_from_excel()


        result['email'] = extract_email(text)
        result['skills'] = extract_skills(text, skills_list)
        result['education'] = extract_education(text, education_keywords)
        result['experience'] = extract_experience(text, experience_keywords)
        result['success'] = True

        logger.info(f"CV parsed successfully - Email: {result['email']}, Skills found: {len(result['skills'].split(',')) if result['skills'] else 0}")

    except Exception as e:
        result['error'] = str(e)
        logger.error(f"CV parsing error: {e}")

    return result