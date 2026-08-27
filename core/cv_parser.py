import os
import re
import time
import logging
import threading

# PDF handling
try:
    import PyPDF2
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import docx
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import pytesseract
    from pdf2image import convert_from_path
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# Excel handling (for the keyword data sheet)
import openpyxl



logger = logging.getLogger('ai_recruitment')


TESSERACT_CMD = r'D:\HR Recruitment System\Tesseract\tesseract.exe'
POPPLER_PATH = r'D:\HR Recruitment System\poppler-26.02.0\Library\bin'

if OCR_AVAILABLE:

    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

    if not os.path.isfile(TESSERACT_CMD):
        logger.error(f"Tesseract executable not found: {TESSERACT_CMD}")
    else:
        try:
            tesseract_version = pytesseract.get_tesseract_version()
            logger.info(f"Tesseract detected: {tesseract_version}")
        except Exception as e:
            logger.error(f"Tesseract validation failed: {e}", exc_info=True)

print(
    f"DEBUG STARTUP: PDFPLUMBER_AVAILABLE={PDFPLUMBER_AVAILABLE}, "
    f"PYPDF2_AVAILABLE={PYPDF2_AVAILABLE}, "
    f"OCR_AVAILABLE={OCR_AVAILABLE}, "
    f"TESSERACT_CMD={TESSERACT_CMD}, "
    f"TESSERACT_EXISTS={os.path.isfile(TESSERACT_CMD)}, "
    f"POPPLER_PATH={POPPLER_PATH}, "
    f"POPPLER_EXISTS={os.path.isdir(POPPLER_PATH)}"
)


VALID_TLDS = {
    'com', 'org', 'net', 'edu', 'gov', 'mil', 'int', 'co', 'io', 'ai',
    'uk', 'us', 'ca', 'au', 'de', 'fr', 'jp', 'cn', 'in', 'lk', 'sg',
    'info', 'biz', 'name', 'pro', 'aero', 'jobs', 'mobi', 'travel',
    'app', 'dev', 'cloud', 'tech', 'online', 'store', 'blog',
}


CV_KEYWORDS_XLSX = os.environ.get(
    'CV_KEYWORDS_XLSX',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cv_keywords.xlsx')
)

try:
    from django.conf import settings as _django_settings
    if getattr(_django_settings, 'CV_KEYWORDS_XLSX', None):
        CV_KEYWORDS_XLSX = str(_django_settings.CV_KEYWORDS_XLSX)
except Exception:
    pass

KEYWORDS_CACHE_TTL = int(os.environ.get('CV_KEYWORDS_CACHE_TTL', '300'))

_keywords_cache = {
    'skills': [],
    'education': [],
    'experience': [],
    'loaded_at': 0.0,
    'source': None,
}
_keywords_lock = threading.Lock()


_FALLBACK_SKILLS = ['python', 'java', 'javascript', 'sql', 'aws', 'docker']
_FALLBACK_EDUCATION = ['bachelor', 'master', 'phd', 'degree', 'university']
_FALLBACK_EXPERIENCE = ['experience', 'worked', 'managed', 'developed']


def _read_single_column_sheet(ws):
    """Read a (Keyword, Active) or (Skill, Category, Active) sheet.

    Returns the list of values in column A where the last column
    ('Active') is not explicitly 'No'/'FALSE'/0. Header row is skipped.
    """
    values = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or row[0] is None:
            continue
        keyword = str(row[0]).strip()
        if not keyword:
            continue

        active_flag = row[-1] if len(row) > 1 else None
        if active_flag is not None:
            active_str = str(active_flag).strip().lower()
            if active_str in ('no', 'false', '0', 'n'):
                continue

        values.append(keyword.lower())
    return values


def load_keywords_from_excel(file_path=None, force_reload=False):
    path = file_path or CV_KEYWORDS_XLSX

    with _keywords_lock:
        cache_is_fresh = (
            not force_reload
            and _keywords_cache['source'] == path
            and (time.time() - _keywords_cache['loaded_at']) < KEYWORDS_CACHE_TTL
            and _keywords_cache['skills']
        )
        if cache_is_fresh:
            return (
                _keywords_cache['skills'],
                _keywords_cache['education'],
                _keywords_cache['experience'],
            )

        if not os.path.exists(path):
            logger.error(f"Keyword workbook not found at {path}. Using fallback keyword lists.")
            skills, education, experience = _FALLBACK_SKILLS, _FALLBACK_EDUCATION, _FALLBACK_EXPERIENCE
        else:
            try:
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

                skills = _read_single_column_sheet(wb['Skills']) if 'Skills' in wb.sheetnames else []
                education = _read_single_column_sheet(wb['EducationKeywords']) if 'EducationKeywords' in wb.sheetnames else []
                experience = _read_single_column_sheet(wb['ExperienceKeywords']) if 'ExperienceKeywords' in wb.sheetnames else []

                wb.close()

                if not skills:
                    logger.warning("Skills sheet was empty; falling back to default skill list.")
                    skills = _FALLBACK_SKILLS
                if not education:
                    education = _FALLBACK_EDUCATION
                if not experience:
                    experience = _FALLBACK_EXPERIENCE

                logger.info(
                    f"Loaded keyword sheet from {path}: "
                    f"{len(skills)} skills, {len(education)} education keywords, "
                    f"{len(experience)} experience keywords"
                )
            except Exception as e:
                logger.error(f"Error reading keyword workbook {path}: {e}. Using fallback keyword lists.")
                skills, education, experience = _FALLBACK_SKILLS, _FALLBACK_EDUCATION, _FALLBACK_EXPERIENCE

        _keywords_cache['skills'] = skills
        _keywords_cache['education'] = education
        _keywords_cache['experience'] = experience
        _keywords_cache['loaded_at'] = time.time()
        _keywords_cache['source'] = path

        return skills, education, experience


# Words that should NEVER be part of a person's name
EXCLUDE_NAME_WORDS = {
    'resume', 'curriculum', 'vitae', 'cv', 'name', 'contact',
    'address', 'phone', 'email', 'linkedin', 'github', 'portfolio',
    'objective', 'summary', 'profile', 'education', 'experience',
    'skills', 'references', 'date', 'page', 'of', 'the', 'and', 'or',
    'to', 'in', 'for', 'with', 'on', 'at', 'from', 'by', 'about',
    'present', 'current', 'january', 'february', 'march', 'april',
    'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december',
    'street', 'avenue', 'road', 'lane', 'drive', 'blvd', 'city', 'state',
    'zip', 'postal', 'area', 'code', 'mobile', 'tel', 'fax', 'www',
    'http', 'https', 'com', 'org', 'net', 'edu', 'gov', 'inc', 'ltd',
    'software', 'engineer', 'developer', 'manager', 'analyst', 'consultant',
    'senior', 'junior', 'lead', 'principal', 'associate', 'executive',
    'professional', 'precis', 'about', 'me', 'personal', 'details',
    'declaration', 'activities', 'achievements', 'awards', 'certificates',
    'referees', 'references', 'trainings', 'projects', 'areas', 'expertise',
    'independent', 'consultant', 'factory', 'accountant', 'assistant',
    'chief', 'operating', 'officer', 'deputy', 'general', 'manager',
    'principle', 'software', 'colonel', 'member', 'member-', 'sri', 'lanka',
    'university', 'college', 'school', 'department', 'group', 'limited',
    'phone', 'email', 'tel', 'fax', 'linkedin', 'skype',
    # Section headers that appear as capitalized words
    'eminent', 'feats', 'qualifications', 'affiliations', 'professional',
    'independent', 'consultant', 'manager', 'process', 'compliance',
    'factory', 'accountant', 'assistant', 'accountant', 'trainings',
    'referees', 'areas', 'expertise', 'projects', 'skills', 'soft',
    'certificates', 'certficates', 'achievements', 'awards', 'activities',
    'personal', 'details', 'declaration', 'references', 'about',
}

# Non-personal email patterns to exclude
EXCLUDE_EMAIL_PATTERNS = [
    'info@', 'support@', 'contact@', 'admin@', 'sales@', 'marketing@',
    'noreply@', 'no-reply@', 'help@', 'careers@', 'jobs@', 'hr@',
    'recruit@', 'enquiries@', 'inquiries@', 'webmaster@', 'postmaster@',
    'abuse@', 'security@', 'billing@', 'accounts@', 'feedback@',
    'service@', 'services@', 'team@', 'hello@', 'press@', 'media@',
]


def _extract_pdf_text_pdfplumber(file_path):
    """Extract text using pdfplumber. Much less prone to inserting spurious
    mid-word spaces than PyPDF2 on Word/LibreOffice-exported PDFs."""
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if page_text:
                text += page_text + "\n"
    return text


def _extract_pdf_text_pypdf2(file_path):
    """Fallback extractor. Known to split words on some fonts (see note
    above _extract_pdf_text_pdfplumber)."""
    text = ""
    with open(file_path, "rb") as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_pdf_text(file_path):
    """Extract text from PDF file with OCR fallback for scanned documents"""
    text = ""

    if PDFPLUMBER_AVAILABLE:
        try:
            text = _extract_pdf_text_pdfplumber(file_path)
            logger.info(f"After pdfplumber: {len(text.strip())} chars")
        except Exception as e:
            logger.error(f"pdfplumber read failed, falling back to PyPDF2: {e}")
            text = ""

    if len(text.strip()) < 100 and PYPDF2_AVAILABLE:
        try:
            pypdf2_text = _extract_pdf_text_pypdf2(file_path)
            if len(pypdf2_text.strip()) > len(text.strip()):
                text = pypdf2_text
            logger.info(f"After PyPDF2 fallback: {len(text.strip())} chars")
        except Exception as e:
            logger.error(f"PyPDF2 read failed: {e}")

    if not PDFPLUMBER_AVAILABLE and not PYPDF2_AVAILABLE:
        logger.warning("Neither pdfplumber nor PyPDF2 available")
        return text

    logger.info(f"After text extraction: {len(text.strip())} chars, OCR_AVAILABLE={OCR_AVAILABLE}, len<100={len(text.strip()) < 100}")

    if len(text.strip()) < 100 and OCR_AVAILABLE:
        try:
            if POPPLER_PATH:
                images = convert_from_path(file_path, poppler_path=POPPLER_PATH)
            else:
                images = convert_from_path(file_path)
            logger.info(f"Converted to {len(images)} image(s) for OCR")

            pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

            if not os.path.isfile(TESSERACT_CMD):
                raise FileNotFoundError(
                    f"Tesseract executable not found: {TESSERACT_CMD}"
                )

            tesseract_version = pytesseract.get_tesseract_version()
            logger.info(
                f"Using Tesseract: {TESSERACT_CMD} | "
                f"Version: {tesseract_version}"
            )

            for page_number, image in enumerate(images, start=1):
                logger.info(f"Running OCR on PDF page {page_number}")
                ocr_text = pytesseract.image_to_string(
                    image,
                    config='--psm 6'
                )
                text += ocr_text + "\n"
                logger.info(
                    f"OCR page {page_number}: {len(ocr_text.strip())} chars"
                )

            logger.info(f"After OCR: {len(text.strip())} chars")
        except Exception as ocr_error:
            logger.error(f"OCR failed: {ocr_error}", exc_info=True)
 
    return text

def extract_docx_text(file_path):
    """Extract text from DOCX file"""
    text = ""
    if not DOCX_AVAILABLE:
        return text
    try:
        doc = docx.Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        logger.error(f"Error reading DOCX {file_path}: {e}")
    return text


def extract_text(file_path):
    """Main text extraction function"""
    if not os.path.exists(file_path):
        return ""
    file_path_lower = file_path.lower()
    if file_path_lower.endswith(".pdf"):
        return extract_pdf_text(file_path)
    elif file_path_lower.endswith(".docx"):
        return extract_docx_text(file_path)
    elif file_path_lower.endswith(".doc"):
        try:
            return extract_docx_text(file_path)
        except:
            return ""
    elif file_path_lower.endswith((".txt", ".rtf")):
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except:
            return ""
    return ""


def preprocess_text(text):

    if not text:
        return text
    
    text = text.replace("mailto:", "")
    text = re.sub(r"\s*@\s*", "@", text)
    

    text = re.sub(r'(\.[a-zA-Z]{2,6})([A-Z0-9])', r'\1 \2', text)

    text = re.sub(r'(\.[a-zA-Z]{2,6})(No\.\d)', r'\1 \2', text)

    text = re.sub(r'[ \t]{2,}', ' ', text)

    return text


def is_personal_email(email):
    """Check if email looks like a personal/candidate email (not a company generic one)"""
    email_lower = email.lower()
    for pattern in EXCLUDE_EMAIL_PATTERNS:
        if pattern in email_lower:
            return False
    local = email_lower.split('@')[0]
    if len(local) <= 2:
        return False
    return True


def clean_email(email):
    """Fix emails where TLD got merged with following text (e.g., gmail.comno -> gmail.com)"""
    email = email.lower().strip()
    parts = email.split('@')
    if len(parts) != 2:
        return email
    domain = parts[1]
    for tld in sorted(VALID_TLDS, key=len, reverse=True):
        tld_with_dot = '.' + tld
        if domain.endswith(tld_with_dot):
            return email
        idx = domain.find(tld_with_dot)
        if idx != -1:
            clean_domain = domain[:idx + len(tld_with_dot)]
            return parts[0] + '@' + clean_domain
    return email


def extract_email(text):
    """Extract the candidate's email from text. Prioritizes emails found early in the document."""
    if not text:
        return ""

    text = preprocess_text(text)

    email_pattern = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,6}\b"
    all_emails = re.findall(email_pattern, text)

    if not all_emails:
        return ""


    seen = set()
    unique_emails = []
    for e in all_emails:
        e_clean = e.lower().strip()
        if e_clean not in seen:
            seen.add(e_clean)
            unique_emails.append(e_clean)

    cleaned_emails = []
    for email in unique_emails:
        email = clean_email(email)
        if any(x in email for x in ['example.com', 'test.com', 'domain.com', 'yourdomain.com']):
            continue
        cleaned_emails.append(email)

    if not cleaned_emails:
        return ""


    header_text = text[:1500].lower()
    for email in cleaned_emails:
        if email in header_text and is_personal_email(email):
            return email

    for email in cleaned_emails:
        if is_personal_email(email):
            return email

    return cleaned_emails[0]


def looks_like_name(words):
    """Check if a list of words looks like a person's name"""
    if not (2 <= len(words) <= 4):
        return False
    

    if not all(w.isalpha() and len(w) > 1 for w in words):
        return False
    

    if any(w.lower() in EXCLUDE_NAME_WORDS for w in words):
        return False
    

    total_len = sum(len(w) for w in words)
    if not (6 <= total_len <= 40):
        return False
    
    all_title_case = all(w[0].isupper() for w in words)
    all_upper = all(w.isupper() for w in words)
    
    if not (all_title_case or all_upper):
        return False
    

    title_indicators = ['engineer', 'developer', 'manager', 'analyst', 'consultant',
                        'director', 'coordinator', 'specialist', 'administrator',
                        'architect', 'designer', 'executive', 'officer', 'president',
                        'vice', 'vp', 'head', 'lead', 'chief', 'senior', 'junior',
                        'process', 'compliance', 'finance', 'quality', 'professional']
    if any(w.lower() in title_indicators for w in words):
        return False
    
    return True


def extract_name_from_text(text, email=""):
    """
    Extract candidate name from CV text.
    Returns (first_name, last_name)
    """
    if not text:
        return "", ""

    lines = [line.strip() for line in text.split('\n') if line.strip()]
    if not lines:
        return "", ""

    # ==================================================================
    # STRATEGY 1: Explicit "Name:" or "Full Name:" labels (first 30 lines)
    # ==================================================================
    name_label_patterns = [
        r'[Nn]ame\s*[:\-]\s*([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+){1,3})',
        r'[Ff]ull\s*[Nn]ame\s*[:\-]\s*([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+){1,3})',
    ]
    header_text = '\n'.join(lines[:30])
    for pattern in name_label_patterns:
        match = re.search(pattern, header_text)
        if match:
            full_name = match.group(1).strip()
            parts = full_name.split()
            filtered = [p for p in parts if p.lower() not in EXCLUDE_NAME_WORDS]
            if len(filtered) >= 2:
                return filtered[0], ' '.join(filtered[1:])
            if len(parts) >= 2:
                return parts[0], ' '.join(parts[1:])

    # ==================================================================
    # STRATEGY 2: First 3 lines - very often the candidate's name is here
    # ==================================================================
    for line_idx in range(min(3, len(lines))):
        line = lines[line_idx]
        # Clean the line - remove common prefixes
        clean_line = re.sub(r'^(?:#|\*|\-|[0-9]+\.)\s*', '', line).strip()
        words = clean_line.split()
        
        if looks_like_name(words):
            # Convert to title case if all uppercase
            if all(w.isupper() for w in words):
                words = [w.title() for w in words]
            return words[0], ' '.join(words[1:])

    # ==================================================================
    # STRATEGY 3: Scan first 15 lines for 2-4 capitalized words
    # ==================================================================
    for line in lines[:15]:
        if len(line) > 60 or len(line) < 4:
            continue
        # Skip lines that are clearly section headers
        if line.startswith('#') or line.startswith('*') or line.startswith('-'):
            continue
        # Skip lines with numbers, special chars, or URLs
        if re.search(r'[0-9@/:;!?&|\-]', line):
            continue

        clean_line = re.sub(r'[^\w\s]', ' ', line)
        words = clean_line.split()

        # Look for consecutive capitalized words
        for j in range(len(words)):
            candidate_words = []
            for k in range(j, min(j + 4, len(words))):
                word = words[k]
                if word and word[0].isupper() and len(word) > 1 and word.isalpha():
                    if word.lower() not in EXCLUDE_NAME_WORDS:
                        candidate_words.append(word)
                    else:
                        break
                elif candidate_words:
                    break

            if len(candidate_words) >= 2:
                full_name = ' '.join(candidate_words)
                if 5 <= len(full_name) <= 40:
                    # Reject if it contains title indicators
                    title_indicators = {'engineer', 'developer', 'manager', 'analyst', 
                                        'consultant', 'director', 'specialist', 'professional',
                                        'process', 'compliance', 'finance', 'quality'}
                    if not any(w.lower() in title_indicators for w in candidate_words):
                        return candidate_words[0], ' '.join(candidate_words[1:])

    # ==================================================================
    # STRATEGY 4: Extract from email local part
    # ==================================================================
    if email and '@' in email:
        local = email.split('@')[0]
        local = re.sub(r'[._\-\d]', ' ', local)
        parts = local.split()
        parts = [p.capitalize() for p in parts if len(p) > 1]
        if len(parts) >= 2:
            return parts[0], ' '.join(parts[1:])
        elif len(parts) == 1:
            return parts[0], ""
    return "", ""


def extract_skills(text, skills_list=None):
    """Extract skills from CV text using the Skills sheet from the workbook"""
    if not text:
        return ""
    if skills_list is None:
        skills_list, _, _ = load_keywords_from_excel()
    text_lower = text.lower()
    found_skills = []
    for skill in skills_list:
        pattern = r'\b' + re.escape(skill) + r'\b'
        if re.search(pattern, text_lower):
            found_skills.append(skill)
    return ", ".join(found_skills) if found_skills else ""


def extract_education(text, education_keywords=None):
    """Extract education information from CV text"""
    if not text:
        return ""
    if education_keywords is None:
        _, education_keywords, _ = load_keywords_from_excel()
    lines = text.split('\n')
    education_lines = []
    for line in lines:
        line_lower = line.lower().strip()
        if any(keyword in line_lower for keyword in education_keywords):
            clean_line = line.strip()
            if len(clean_line) > 10:
                education_lines.append(clean_line)
    return "\n".join(education_lines[:10]) if education_lines else ""


def extract_experience(text, experience_keywords=None):
    """Extract work experience from CV text"""
    if not text:
        return ""
    if experience_keywords is None:
        _, _, experience_keywords = load_keywords_from_excel()
    lines = text.split('\n')
    experience_lines = []
    for line in lines:
        line_lower = line.lower().strip()
        if any(keyword in line_lower for keyword in experience_keywords):
            clean_line = line.strip()
            if len(clean_line) > 10:
                experience_lines.append(clean_line)
    return "\n".join(experience_lines[:15]) if experience_lines else ""


def parse_cv(file_path):
    """Complete CV parsing - extracts all information"""
    result = {
        'text': '',
        'email': '',
        'first_name': '',
        'last_name': '',
        'skills': '',
        'education': '',
        'experience': '',
        'success': False,
        'error': ''
    }

    try:
        text = extract_text(file_path)
        result['text'] = text

        if not text.strip():
            result['error'] = 'No text could be extracted from the file'
            return result

        skills_list, education_keywords, experience_keywords = load_keywords_from_excel()


        result['email'] = extract_email(text)
        first_name, last_name = extract_name_from_text(text, result['email'])
        result['first_name'] = first_name
        result['last_name'] = last_name
        result['skills'] = extract_skills(text, skills_list)
        result['education'] = extract_education(text, education_keywords)
        result['experience'] = extract_experience(text, experience_keywords)
        result['success'] = True

        logger.info(
            f"CV parsed - Name: '{first_name} {last_name}', "
            f"Email: '{result['email']}'"
        )

    except Exception as e:
        result['error'] = str(e)
        logger.error(f"CV parsing error: {e}")

    return result